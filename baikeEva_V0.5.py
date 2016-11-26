# baikeEva V 0.1
# author: babacheku


from openpyxl import load_workbook
import baikevalib

wb = load_workbook("test_1.xlsx")

for wsIndex in range(len(wb.sheetnames)):

    # 打开worksheet
    ws = wb.worksheets[wsIndex]
    # 插入表头并获取词条列数
    entryColumnIndex = baikevalib.sheet_header_insert(ws)

    # 判断表格格式是否正确
    if entryColumnIndex == -1:
        print("表格格式有误，请检查后重试！")
        continue
    else:
        print("表头插入成功")

    # sheet内部循环
    for rowIndex in range(3, (ws.max_row + 1)):

        # 待查询词条名称
        keyWord = ws.cell(row=rowIndex, column=entryColumnIndex).value.lstrip().rstrip()
        print("当前词条：" + keyWord)

        # 生成词条网址
        baikePageSoup = baikevalib.get_page_soup("http://baike.baidu.com/item/", keyWord)

        # 判断词条是被百度百科直接收录
        inBaiduBaike = baikevalib.is_baidu_baike(baikePageSoup)

        # 获取合法百科页面数据
        if inBaiduBaike is False:

            # 百度百科未直接收录
            print("百度未直接收录")

            # 获取合法百科网址
            baikeURL = baikevalib.get_legal_page_url("http://www.baidu.com/s?wd=", keyWord)

            if baikeURL is None:
                print("百度未收录")
                continue
            else:
                print("百度百科间接收录")
                baikePageSoup = baikevalib.get_page_soup(baikeURL)
        else:
            print("百度百科直接收录")
            baikeURL = "http://baike.baidu.com/item/" + keyWord

        # 判断词条是否为多义词
        polysemousTag = baikePageSoup.find_all("div", {"class": {"lemmaWgt-subLemmaListTitle",
                                                                 "polysemantList-header-title"}})
        if len(polysemousTag) > 0:
            print("词条为多义词")
            isPolysemousEntry = True
            # 词条是多义词，获取合法义项页面
            baikeURL = baikevalib.get_legal_item_page_url(baikePageSoup, baikevalib.KeyList.Aero_Avia_Keys)
            baikePageSoup = baikevalib.get_page_soup(baikeURL)
            # 获取义项编码
            itemCode = baikevalib.get_item_code(baikePageSoup)
        else:
            print("词条非多义词")
            isPolysemousEntry = False
            itemCode = None

        print("词条页面地址为" + baikeURL)
        # 获取词条页面数据
        # 获取词条名称
        entryTitle = baikePageSoup.find("h1").get_text()
        print("百科词条名称为：" + entryTitle)

        # 是否是"特色词条"
        excellentTag = baikePageSoup.find("a", {"class": "posterFlag excellent-icon"})
        if excellentTag is not None:
            excellentIcon = True
            print("是特色词条")
        else:
            excellentIcon = False
            print("非特色词条")

        # 是否被"科普中国百科"收录
        inKepuBaike = baikevalib.is_kepu_baike(baikePageSoup)
        if inKepuBaike:
            print("科普百科收录")
        else:
            print("科普百科未收录")

        if inKepuBaike or excellentIcon:
            baikevalib.excel_data_import(ws, rowIndex, entryTitle, inBaiduBaike, inKepuBaike, excellentIcon, baikeURL,
                                         isPolysemousEntry, itemCode)
            continue
        else:
            pass

        # 获取概述字数
        summarySum = baikevalib.get_summary_sum(baikePageSoup)
        print("概述共有%d个字" % summarySum)

        # 获取基本信息栏条数
        basicInfoSum = len(baikePageSoup.find_all("dt", {"class": "basicInfo-item name"}))
        print("共有%d条基本信息栏" % basicInfoSum)

        # 获取一级目录条数
        catalog1Sum = len(baikePageSoup.find_all("li", {"class": "level1"}))
        print("共有%d条一级目录" % catalog1Sum)

        # 获取正文字数
        contentSum = baikevalib.get_content_sum(baikePageSoup)
        print("正文共有%d个字" % contentSum)

        # 获取参考文献条数
        referenceSum = len(baikePageSoup.find_all("li", {"class": {"reference-item ", "reference-item more"}}))
        print("共有%d条参考文献" % referenceSum)

        # 获取图片张数
        pictureSum = baikevalib.get_picture_sum(baikePageSoup)
        print("共有%d张图片" % pictureSum)

        # 词条等级评定
        entryRank = baikevalib.entry_evaluate(summarySum, basicInfoSum, catalog1Sum, contentSum, referenceSum,
                                              pictureSum)

        baikevalib.excel_data_import(ws, rowIndex, entryTitle, inBaiduBaike, inKepuBaike, excellentIcon, baikeURL,
                                     isPolysemousEntry, itemCode, entryRank, summarySum, basicInfoSum, catalog1Sum,
                                     contentSum, referenceSum, pictureSum)

        wb.save("test_1.xlsx")
